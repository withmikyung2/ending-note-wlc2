import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E5F", end_color="1F4E5F", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
EXAMPLE_FONT = Font(name=FONT_NAME, italic=True, color="7F7F7F", size=10)
BODY_FONT = Font(name=FONT_NAME, size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E5F")
SUB_FONT = Font(name=FONT_NAME, size=10, color="595959")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)


def style_header(ws, ncols, row=1, height=22):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_example_row(ws, row, values):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = EXAMPLE_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_validation(ws, col_letter, options, row_start=3, row_end=500):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")


# ---------------------------------------------------------------------------
# 00_안내
# ---------------------------------------------------------------------------
ws = wb.create_sheet("00_안내")
ws.sheet_view.showGridLines = False
set_widths(ws, [110])
ws["A1"] = "GYA_AMS 데이터 시트 (Draft v0.1)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "GYA · GWC 뚬놉(Tom Nop) 지역 교육혁신 네트워크 — 학사관리 데이터 템플릿"
ws["A2"].font = SUB_FONT

lines = [
    "",
    "■ 이 파일의 목적",
    "이 파일은 GYA Center와 뚬놉(Tom Nop) 지역 약 7개 Community Learning Room의 학생·교사·출석·평가·활동을",
    "하나의 표로 통합 관리하기 위한 시작용(MVP) 데이터베이스입니다. Google Drive에 업로드해 Google Sheets로 열어",
    "GYA Hub 스태프와 GWC 본부가 함께 편집·조회합니다.",
    "",
    "■ 탭 구성",
    "01_거점        Community Learning Room(공부방) 목록",
    "02_학생        학생 등록 정보",
    "03_교사        Local Teacher 등록 정보",
    "04_출석        회차별 출석 기록",
    "05_평가_EKM    English·Khmer·Math 기초학습 baseline/중간/기말 평가",
    "06_교사연수    월례회의·코칭·집중연수 등 교사 전문성 개발 기록",
    "07_활동로그    방문·물품지원·행사 등 현장 활동 기록",
    "08_대시보드    자동 집계되는 핵심 지표(KPI) 요약",
    "",
    "■ 색 안내",
    "노란색 배경 셀 = 직접 입력하는 칸입니다.",
    "회색 기울임 글씨 행 = 작성 예시입니다. 실제 데이터를 입력할 때는 이 행 아래에 새 행을 추가하세요.",
    "08_대시보드의 숫자는 수식으로 자동 계산되며 직접 입력하지 않습니다.",
    "",
    "■ ID 규칙",
    "거점ID 예: TN-01 ~ TN-07 (Tom Nop 1~7번 공부방)",
    "학생ID 예: TN01-S001 (거점코드-S+일련번호)",
    "교사ID 예: TN01-T01 (거점코드-T+일련번호)",
    "",
    "■ 다음 단계 (착수 100일 계획과 연결)",
    "1) 01_거점 탭에 7개 공부방 현황 실사 결과를 입력합니다.",
    "2) 02_학생 / 03_교사 탭에 baseline 등록 정보를 입력합니다.",
    "3) 04_출석 탭은 Google Forms 응답이 자동으로 쌓이도록 연동하는 것을 권장합니다 (별도 안내 문서 docs/forms-spec.md 참고).",
    "4) 05_평가_EKM 탭에 진단평가(baseline) 결과를 입력해 기초학습 성장을 추적합니다.",
    "",
    "이 파일은 Draft이며, 실제 운영 데이터가 쌓이면 GYA Hub 운영팀과 협의해 항목을 조정합니다.",
]
r = 4
for line in lines:
    cell = ws.cell(row=r, column=1, value=line)
    if line.startswith("■"):
        cell.font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E5F")
    else:
        cell.font = BODY_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ---------------------------------------------------------------------------
# 01_거점 (Learning Rooms)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("01_거점")
headers = ["거점ID", "거점명", "마을(Village)", "위치 설명", "담당 교사ID", "개설일",
           "등록 학생수(자동)", "상태", "비고"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, len(headers))
set_widths(ws, [10, 20, 16, 26, 12, 12, 14, 10, 24])
ws.freeze_panes = "A2"

add_example_row(ws, 2, [
    "TN-01", "뚬놉 1번 공부방", "Tom Nop", "마을회관 옆 개방형 학습공간",
    "TN01-T01", "2024-03-01",
    '=COUNTIFS(\'02_학생\'!$F$2:$F$1000,$A2,\'02_학생\'!$L$2:$L$1000,"활동")',
    "운영중", "예시 행 — 삭제하지 말고 아래에 새 행 추가"
])
for row in range(3, 10):
    ws.cell(row=row, column=1, value=f"TN-0{row-1}" if row-1 <= 7 else "")
    ws.cell(row=row, column=7,
             value=f'=COUNTIFS(\'02_학생\'!$F$2:$F$1000,$A{row},\'02_학생\'!$L$2:$L$1000,"활동")')
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = BORDER
        ws.cell(row=row, column=c).font = BODY_FONT
        ws.cell(row=row, column=c).fill = INPUT_FILL
add_validation(ws, "H", ["운영중", "준비중", "일시중단", "종료"])

# ---------------------------------------------------------------------------
# 02_학생 (Students)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("02_학생")
headers = ["학생ID", "이름", "성별", "생년월일", "나이(자동)", "거점ID", "학교", "학년",
           "등록일", "보호자 연락처", "사진동의여부", "상태"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, len(headers))
set_widths(ws, [12, 14, 8, 12, 10, 10, 16, 8, 12, 16, 12, 10])
ws.freeze_panes = "A2"

add_example_row(ws, 2, [
    "TN01-S001", "Sok Dara", "여", "2015-05-12", "=DATEDIF(D2,TODAY(),\"y\")",
    "TN-01", "Tom Nop Primary School", "4", "2024-03-01", "+855-XX-XXX-XXX", "Y", "활동"
])
for row in range(3, 12):
    ws.cell(row=row, column=5, value=f"=IF(D{row}=\"\",\"\",DATEDIF(D{row},TODAY(),\"y\"))")
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = BORDER
        ws.cell(row=row, column=c).font = BODY_FONT
        ws.cell(row=row, column=c).fill = INPUT_FILL
add_validation(ws, "C", ["남", "여"])
add_validation(ws, "K", ["Y", "N"])
add_validation(ws, "L", ["활동", "휴학", "졸업", "중도탈락"])

# ---------------------------------------------------------------------------
# 03_교사 (Local Teachers)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("03_교사")
headers = ["교사ID", "이름", "거점ID", "연락처", "담당 과목/영역", "채용일",
           "행동강령 서명일", "상태", "비고"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, len(headers))
set_widths(ws, [10, 14, 10, 16, 20, 12, 14, 10, 20])
ws.freeze_panes = "A2"

add_example_row(ws, 2, [
    "TN01-T01", "Chan Sopheak", "TN-01", "+855-XX-XXX-XXX",
    "English / Reading Club", "2024-02-15", "2024-02-20", "활동", "예시 행"
])
for row in range(3, 12):
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = BORDER
        ws.cell(row=row, column=c).font = BODY_FONT
        ws.cell(row=row, column=c).fill = INPUT_FILL
add_validation(ws, "H", ["활동", "휴직", "퇴직"])

# ---------------------------------------------------------------------------
# 04_출석 (Attendance)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("04_출석")
headers = ["날짜", "거점ID", "학생ID", "학생이름(자동)", "출석여부", "비고"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, len(headers))
set_widths(ws, [12, 10, 12, 16, 10, 24])
ws.freeze_panes = "A2"

add_example_row(ws, 2, [
    "2026-09-01", "TN-01", "TN01-S001",
    '=IFERROR(INDEX(\'02_학생\'!$B:$B,MATCH(C2,\'02_학생\'!$A:$A,0)),"")',
    "출석", "예시 행 — Google Forms 응답을 이 아래에 자동 연동하는 것을 권장"
])
for row in range(3, 30):
    ws.cell(row=row, column=4,
             value=f'=IFERROR(INDEX(\'02_학생\'!$B:$B,MATCH(C{row},\'02_학생\'!$A:$A,0)),"")')
    for c in [1, 2, 3, 5, 6]:
        ws.cell(row=row, column=c).border = BORDER
        ws.cell(row=row, column=c).font = BODY_FONT
        ws.cell(row=row, column=c).fill = INPUT_FILL
add_validation(ws, "E", ["출석", "결석", "지각"])

# ---------------------------------------------------------------------------
# 05_평가_EKM (Assessment)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("05_평가_EKM")
headers = ["학생ID", "학생이름(자동)", "평가유형", "영역", "점수/레벨(1-3)", "평가일", "평가자", "비고"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, len(headers))
set_widths(ws, [12, 16, 10, 10, 14, 12, 12, 20])
ws.freeze_panes = "A2"

add_example_row(ws, 2, [
    "TN01-S001",
    '=IFERROR(INDEX(\'02_학생\'!$B:$B,MATCH(A2,\'02_학생\'!$A:$A,0)),"")',
    "baseline", "English", "1", "2026-09-01", "GYA Hub", "예시 행"
])
for row in range(3, 20):
    ws.cell(row=row, column=2,
             value=f'=IFERROR(INDEX(\'02_학생\'!$B:$B,MATCH(A{row},\'02_학생\'!$A:$A,0)),"")')
    for c in [1, 3, 4, 5, 6, 7, 8]:
        ws.cell(row=row, column=c).border = BORDER
        ws.cell(row=row, column=c).font = BODY_FONT
        ws.cell(row=row, column=c).fill = INPUT_FILL
add_validation(ws, "C", ["baseline", "중간평가", "기말평가"])
add_validation(ws, "D", ["English", "Khmer", "Math"])
add_validation(ws, "E", ["1", "2", "3"])

# ---------------------------------------------------------------------------
# 06_교사연수 (Teacher Development)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("06_교사연수")
headers = ["날짜", "교사ID", "교사이름(자동)", "연수유형", "내용", "비고"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, len(headers))
set_widths(ws, [12, 10, 14, 16, 36, 20])
ws.freeze_panes = "A2"

add_example_row(ws, 2, [
    "2026-08-05", "TN01-T01",
    '=IFERROR(INDEX(\'03_교사\'!$B:$B,MATCH(B2,\'03_교사\'!$A:$A,0)),"")',
    "월례 Teacher Community Meeting", "학생사례 공유, 8월 수업안 점검", "예시 행"
])
for row in range(3, 15):
    ws.cell(row=row, column=3,
             value=f'=IFERROR(INDEX(\'03_교사\'!$B:$B,MATCH(B{row},\'03_교사\'!$A:$A,0)),"")')
    for c in [1, 2, 4, 5, 6]:
        ws.cell(row=row, column=c).border = BORDER
        ws.cell(row=row, column=c).font = BODY_FONT
        ws.cell(row=row, column=c).fill = INPUT_FILL
add_validation(ws, "D", ["월례회의", "순회코칭", "집중연수", "Lesson Study", "연간 Review"])

# ---------------------------------------------------------------------------
# 07_활동로그 (Activity Log) -- pre-filled with real Tom Nop visit (news article)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("07_활동로그")
headers = ["날짜", "거점ID", "활동유형", "내용", "지원물품", "참여인원", "작성자", "비고"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, len(headers))
set_widths(ws, [12, 10, 12, 34, 20, 10, 10, 18])
ws.freeze_panes = "A2"

real_rows = [
    ["2026-07-23", "TN-01", "현장방문", "GWC 대표 이미경, 뚬놉 지역 공부방 방문 및 시설·필요물품 확인", "", "", "Cindy", "실제 기록"],
    ["2026-07-25", "TN-01", "물품지원", "공부방 4곳에 화이트보드 4개, 공책 200권 전달", "화이트보드 4, 공책 200권", "", "Cindy", "실제 기록"],
]
for i, vals in enumerate(real_rows, start=2):
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.fill = INPUT_FILL
for row in range(4, 15):
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = BORDER
        ws.cell(row=row, column=c).font = BODY_FONT
        ws.cell(row=row, column=c).fill = INPUT_FILL
add_validation(ws, "C", ["현장방문", "물품지원", "행사", "교육", "기타"])

# ---------------------------------------------------------------------------
# 08_대시보드 (Dashboard)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("08_대시보드")
ws.sheet_view.showGridLines = False
set_widths(ws, [30, 16, 44])
ws["A1"] = "GYA_AMS 핵심 지표 (자동 집계)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "아래 숫자는 수식으로 계산되며 직접 입력하지 않습니다."
ws["A2"].font = SUB_FONT

kpi_rows = [
    ("운영중 거점 수", '=COUNTIFS(\'01_거점\'!$H$2:$H$1000,"운영중")', "01_거점 상태='운영중' 개수"),
    ("등록 학생 수(활동)", '=COUNTIFS(\'02_학생\'!$L$2:$L$1000,"활동")', "02_학생 상태='활동' 개수"),
    ("등록 Local Teacher 수(활동)", '=COUNTIFS(\'03_교사\'!$H$2:$H$1000,"활동")', "03_교사 상태='활동' 개수"),
    ("전체 출석률", '=IFERROR(COUNTIFS(\'04_출석\'!$E$2:$E$1000,"출석")/COUNTIFS(\'04_출석\'!$E$2:$E$1000,"<>"),"데이터 없음")', "출석 건수 / 전체 출석기록 건수"),
    ("Baseline 평가 등록 건수", '=COUNTIFS(\'05_평가_EKM\'!$C$2:$C$1000,"baseline")', "05_평가_EKM 평가유형='baseline'"),
    ("이번 분기 교사연수 실시 건수", '=COUNTIFS(\'06_교사연수\'!$A$2:$A$1000,">="&EOMONTH(TODAY(),-3)+1)', "최근 3개월간 06_교사연수 기록 수"),
    ("최근 30일 활동로그 건수", '=COUNTIFS(\'07_활동로그\'!$A$2:$A$1000,">="&TODAY()-30)', "최근 30일간 07_활동로그 기록 수"),
]
r = 4
ws.cell(row=r, column=1, value="지표").font = HEADER_FONT
ws.cell(row=r, column=2, value="값").font = HEADER_FONT
ws.cell(row=r, column=3, value="산출 기준").font = HEADER_FONT
for c in range(1, 4):
    ws.cell(row=r, column=c).fill = HEADER_FILL
r += 1
for name, formula, note in kpi_rows:
    ws.cell(row=r, column=1, value=name).font = BODY_FONT
    vcell = ws.cell(row=r, column=2, value=formula)
    vcell.font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E5F")
    vcell.number_format = '0.0%' if "출석률" in name else '0'
    ws.cell(row=r, column=3, value=note).font = SUB_FONT
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = BORDER
    r += 1

wb.active = wb["00_안내"]
wb.save("gya_ams_data_v0.1.xlsx")
print("saved")
